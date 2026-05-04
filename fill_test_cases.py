from openpyxl import load_workbook

DATA = [
    ("Neg_001", "S", "uba kohomda bn", "ඔබ කොහොමද බන්"),
    ("Neg_002", "S", "mta enna ba", "මට එන්න බැහැ"),
    ("Neg_003", "S", "oya kohed inne", "ඔබ කොහෙද ඉන්නේ"),
    ("Neg_004", "S", "ubata mokda", "ඔබට මොකද්ද"),
    ("Neg_005", "S", "mama awa", "මම පැමිණියෙමි"),
    ("Neg_006", "S", "oya hithanne mokda", "ඔබ සිතන්නේ කුමක්ද"),
    ("Neg_007", "S", "api yamu", "අපි යමුද"),
    ("Neg_008", "S", "ubata pissuda 😂", "ඔබට පිස්සුද 😂"),
    ("Neg_009", "S", "mta therune na", "මට තේරුණේ නැත"),
    ("Neg_010", "S", "oya kiyanne mokda", "ඔබ කියන්නේ කුමක්ද"),
    ("Neg_011", "M", "mama gedara enawa", "මම ගෙදර පැමිණෙමින් සිටිමි"),
    ("Neg_012", "M", "ubala mokda karanne", "ඔබලා මොනවා කරමින් සිටිනවාද"),
    ("Neg_013", "M", "api match ekak gahanawa", "අපි තරඟයක් ක්‍රීඩා කරමින් සිටිමු"),
    ("Neg_014", "M", "mama kiwwa ne ehema epa kiyala", "මම කිව්වා නේ එහෙම කරන්න එපා කියලා"),
    ("Neg_015", "M", "oya hithan inne mokda", "ඔබ සිතමින් සිටින්නේ කුමක්ද"),
    ("Neg_016", "M", "mama office eke inne", "මම කාර්යාලයේ සිටිමි"),
    ("Neg_017", "M", "api trip ekak yanna hithan inne", "අපි සංචාරයක් යාමට සිතමින් සිටිමු"),
    ("Neg_018", "M", "oya kohomda karanne", "ඔබ මෙය කෙසේ කරන්නේද"),
    ("Neg_019", "M", "mama call karanna giya", "මම දුරකථනය කළෙමි"),
    ("Neg_020", "M", "ubata puluwanda meka karanna", "ඔබට මෙය කළ හැකිද"),
    ("Neg_021", "L", "mama ada ude nagitala class giya", "මම අද උදේ නැගිටලා පන්ති ගියෙමි"),
    ("Neg_022", "L", "api trip ekak plan karanawa", "අපි සංචාරයක් සැලසුම් කරමින් සිටිමු"),
    ("Neg_023", "L", "mama oyata message ekak damma", "මම ඔබට පණිවිඩයක් යවා ඇත"),
    ("Neg_024", "L", "ubata therunaganna amarui", "ඔබට තේරුම් ගැනීමට අපහසුය"),
    ("Neg_025", "L", "api weekend eke beach yanna plan karala inne", "අපි සති අන්තයේ වෙරළට යාමට සැලසුම් කරමින් සිටිමු"),
    ("Neg_026", "S", "oya enne kawadda", "ඔබ එන්නේ කවදාද"),
    ("Neg_027", "S", "mta eka ba", "මට එය කළ නොහැක"),
    ("Neg_028", "S", "ubata kiwwa ne", "මම ඔබට කියා ඇත"),
    ("Neg_029", "S", "api enawa tikak passe", "අපි ටික වේලාවකින් පැමිණෙමු"),
    ("Neg_030", "S", "oya mokda kiyanne", "ඔබ කුමක් පවසන්නේද"),
    ("Neg_031", "M", "mama bus eke inne", "මම බස් රථයේ සිටිමි"),
    ("Neg_032", "M", "api movie ekak balanna yanne", "අපි චිත්‍රපටයක් නැරඹීමට යමු"),
    ("Neg_033", "M", "oya busy da", "ඔබ කාර්යබහුලද"),
    ("Neg_034", "M", "mama link ekak ewwa", "මම සබැඳියක් යවා ඇත"),
    ("Neg_035", "M", "ubata meka karanna puluwanda", "ඔබට මෙය කළ හැකිද"),
    ("Neg_036", "M", "mama class eke inne", "මම පන්තියේ සිටිමි"),
    ("Neg_037", "M", "api meeting ekata yanne", "අපි රැස්වීමට යමින් සිටිමු"),
    ("Neg_038", "M", "oya phone eka busy nisa katha karanna ba", "ඔබගේ දුරකථනය කාර්යබහුල බැවින් කතා කළ නොහැක"),
    ("Neg_039", "M", "mama reply ekak denna hithan inne", "මම පිළිතුරක් ලබාදීමට සිතමින් සිටිමි"),
    ("Neg_040", "M", "ubata help ekak one", "ඔබට උදව්වක් අවශ්‍යද"),
    ("Neg_041", "L", "mama office giya passe busy una", "මම කාර්යාලයට ගොස් පසුව කාර්යබහුල වුණෙමි"),
    ("Neg_042", "L", "api project ekak karanawa team ekka", "අපි කණ්ඩායමක් සමඟ ව්‍යාපෘතියක් කරමින් සිටිමු"),
    ("Neg_043", "L", "mama message damma busy nisa reply karanna ba", "මම පණිවිඩයක් යවා ඇත නමුත් පිළිතුරු දීමට නොහැකි විය"),
    ("Neg_044", "L", "ubata explain karanna puluwan", "ඔබට පැහැදිලි කර දිය හැක"),
    ("Neg_045", "L", "api trip plan karala inne", "අපි සංචාරයක් සැලසුම් කරමින් සිටිමු"),
    ("Neg_046", "S", "oya mokakda kiyanne", "ඔබ කුමක් පවසන්නේද"),
    ("Neg_047", "S", "mta therenne na", "මට තේරෙන්නේ නැත"),
    ("Neg_048", "S", "ubata mokakda wela thiyenne", "ඔබට කුමක් සිදුවී තිබේද"),
    ("Neg_049", "S", "api kohomada yanne", "අපි කෙසේ යමුද"),
    ("Neg_050", "S", "oya hithanne mokakda", "ඔබ සිතන්නේ කුමක්ද"),
]


def main() -> None:
    workbook_path = "Assignment 1 - Test cases.xlsx"
    wb = load_workbook(workbook_path)
    ws = wb[wb.sheetnames[0]]

    # Fill only the required columns: A-D
    for i, (tc_id, length_type, inp, expected) in enumerate(DATA, start=2):
        ws.cell(row=i, column=1, value=tc_id)
        ws.cell(row=i, column=2, value=length_type)
        ws.cell(row=i, column=3, value=inp)
        ws.cell(row=i, column=4, value=expected)

    wb.save(workbook_path)
    print(f"Wrote {len(DATA)} test cases to {workbook_path} (columns A-D only).")


if __name__ == "__main__":
    main()
